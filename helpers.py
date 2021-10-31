#helpers
import pandas as pd
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

def getExcel(import_file_path, sheet=0, filterms2=True):
    wb = openpyxl.load_workbook(filename = import_file_path, read_only=True)
    sheet = wb.worksheets[sheet]
    crn = getCorners(sheet)
    try:
        df = pd.DataFrame()
        df = pd.read_excel (import_file_path, header = 6, usecols = range(0, crn[1]))
        df.rename(columns=lambda x: x.replace(".", "_"), inplace = True)
        if (filterms2): 
            df = df.loc[df['MS_MS_spectrum'].notnull()]
        #df = df.loc[df['Flag'].isnull()]
        return df
    except:
        raise
    else:
        print(df.columns)
        print(df.shape)

def mergeSheets(metadata, cnumber, merged):
    header = openpyxl.Workbook()    
    for r in range(1, metadata.active.max_row+1):
        for c in range(1, metadata.active.max_column+1):
            header.active.cell(row = r, column = c, value = metadata.active.cell(r,c).value)
    header.active.insert_cols(1, cnumber)
    
    for r in dataframe_to_rows(merged, index=False, header=True):
        header.active.append(r)
    return header

def getDataExcel(path, sheet):
    try:
        #import_file_path = os.path.join(workingfolder, filename)
        crn = getCorners(sheet)
        df = pd.read_excel (path, header = crn[0], usecols = range(crn[1],crn[3]))
        df.rename(columns=lambda x: x.replace(".", "_"), inplace = True)
        df.index = df.iloc[:, 0]
    except:
        raise
    return df

def getTgtExcel(path, sheet):
    try:
        #import_file_path = os.path.join(workingfolder, filename)
        crn = getCorners(sheet)
        df = pd.read_excel (path, header = crn[0], usecols = range(0, crn[1]))
        df.rename(columns=lambda x: x.replace(".", "_"), inplace = True)
        df.index = df["Feature_ID"]
    except:
        raise
    return df

def getrExcel(path):
    try:
        #import_file_path = os.path.join(workingfolder, filename)
        df = pd.read_csv(path, delimiter='\t')
        df.rename(columns=lambda x: x.replace(".", "_"), inplace = True)
        df.index = df["File name"]
        df = df.iloc[:, range(7,28)]
    except:
        raise
    return df

def getCorners(sheet):
    corner = [0,0,0,0]
    corner[2] = sheet.max_row
    corner[3] = sheet.max_column
    for i in range(1, sheet.max_row):
        corner[0] = i-1;
        if (sheet.cell(i,1).value != None):
            break
    for i in range(1, sheet.max_column):
        corner[1] = i-1;
        if (sheet.cell(1,i).value != None):
            break
    return corner

def makeMetadata(sheet):
    metadata = openpyxl.Workbook()
    crnr = getCorners(sheet)
    md = metadata.active
    rr = 1
    for r in range(1, crnr[0]+1):
        cc = 1
        for c in range(crnr[1]+1, crnr[3]+1):
            value = sheet.cell(r, c).value
            md.cell(rr, cc,).value = value
            cc += 1
        rr += 1
    return metadata
